"""
Serviço de exportação Excel — reutilizável em qualquer router.

Princípios LGPD aplicados:
- CPF/CNPJ é mascarado por padrão (apenas últimos 2 dígitos visíveis).
- Campos de contato pessoal (email, endereço) são omitidos do relatório de OS.
- Nenhum dado sensível é persistido em disco — o arquivo é gerado em memória.
"""
import io
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi.responses import StreamingResponse


def _mascarar_cpf(valor: str | None) -> str:
    """Exibe apenas os 2 últimos caracteres: ex. '***.***.***-00'."""
    if not valor:
        return "-"
    limpo = "".join(c for c in valor if c.isdigit())
    return "*" * (len(limpo) - 2) + limpo[-2:] if len(limpo) >= 2 else "***"


def gerar_excel(
    dados: list[dict[str, Any]],
    nome_aba: str = "Dados",
    nome_arquivo: str = "relatorio.xlsx",
) -> StreamingResponse:
    """
    Recebe uma lista de dicionários e retorna um StreamingResponse com o .xlsx.
    Não salva nada em disco — tudo gerado em memória (io.BytesIO).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = nome_aba

    if not dados:
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"},
        )

    colunas = list(dados[0].keys())

    # Cabeçalho estilizado
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, nome_col in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nome_col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for row_idx, linha in enumerate(dados, start=2):
        for col_idx, chave in enumerate(colunas, start=1):
            ws.cell(row=row_idx, column=col_idx, value=linha.get(chave, ""))

    # Ajusta largura das colunas automaticamente
    for col_idx, nome_col in enumerate(colunas, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = len(str(nome_col))
        for row_idx in range(2, len(dados) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(val or "")))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"},
    )


def exportar_clientes(clientes: list) -> StreamingResponse:
    """
    Exporta lista de clientes com CPF/CNPJ mascarado (LGPD).
    Email e endereço são incluídos apenas se o relatório for interno.
    """
    dados = [
        {
            "ID": c.id,
            "Nome": c.nome,
            "CPF/CNPJ (mascarado)": _mascarar_cpf(c.cpf_cnpj),
            "Telefone": c.telefone or "-",
        }
        for c in clientes
    ]
    return gerar_excel(dados, nome_aba="Clientes", nome_arquivo="clientes.xlsx")


def exportar_ordens_servico(ordens: list) -> StreamingResponse:
    """
    Exporta ordens de serviço sem dados pessoais do cliente (LGPD).
    O vínculo com o cliente é feito pelo ID, não pelo nome.
    """
    dados = [
        {
            "OS": os.id,
            "Cliente ID": os.cliente_id,
            "Status": os.status,
            "Abertura": os.data_abertura.strftime("%d/%m/%Y") if os.data_abertura else "-",
            "Previsao": os.data_previsao.strftime("%d/%m/%Y") if os.data_previsao else "-",
            "Problema": os.descricao_problema or "-",
            "Total (R$)": f"{os.valor_total:.2f}",
        }
        for os in ordens
    ]
    return gerar_excel(dados, nome_aba="Ordens de Servico", nome_arquivo="ordens_servico.xlsx")


def exportar_financeiro(movimentacoes: list) -> StreamingResponse:
    """Exporta movimentações financeiras."""
    dados = [
        {
            "ID": m.id,
            "Tipo": m.tipo,
            "Valor (R$)": f"{m.valor:.2f}",
            "Descricao": m.descricao or "-",
            "Categoria": m.categoria or "-",
            "Data": m.data.strftime("%d/%m/%Y") if m.data else "-",
            "OS Vinculada": m.os_id or "-",
        }
        for m in movimentacoes
    ]
    return gerar_excel(dados, nome_aba="Financeiro", nome_arquivo="financeiro.xlsx")
